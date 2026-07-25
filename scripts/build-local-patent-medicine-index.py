#!/usr/bin/env python3
"""Build a compact outpatient Chinese patent-medicine label index from supplied data."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = PROJECT_ROOT / "药学基础数据/药品说明书数据库_医药数据查询/药品详细信息_总.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "src/data/local-patent-medicine-index.json"


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean(value: object, limit: int) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()[:limit]


def generic_name(value: object) -> str:
    return re.sub(r"[（(][^）)]*[）)]", "", clean(value, 160)).strip()


def fingerprint(record: dict[str, str]) -> str:
    serialized = json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return "sha256:" + hashlib.sha256(serialized.encode()).hexdigest()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    source = args.source.resolve()
    output = args.output.resolve()
    if not source.is_file():
        raise SystemExit(f"Local medicine workbook not found: {source}")

    worksheet = load_workbook(source, read_only=True, data_only=True).active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    required = [
        "产品名称", "包装规格", "批准文号", "生产厂家", "药品类型", "主分类", "子分类",
        "详情链接", "功能主治/适应症", "用法用量", "不良反应", "禁忌", "注意事项",
        "儿童用药", "老年患者用药", "孕妇及哺乳期妇女用药", "药物相互作用",
    ]
    missing = [name for name in required if name not in headers]
    if missing:
        raise SystemExit(f"Local medicine workbook missing columns: {', '.join(missing)}")
    indexes = {name: headers.index(name) for name in required}

    entries: dict[str, dict[str, str]] = {}
    source_row_count = 0
    for row in rows:
        source_row_count += 1
        if clean(row[indexes["主分类"]], 20) != "中成药" or clean(row[indexes["药品类型"]], 20) != "非处方药":
            continue
        name = generic_name(row[indexes["产品名称"]])
        indication = clean(row[indexes["功能主治/适应症"]], 700)
        approval = clean(row[indexes["批准文号"]], 100)
        manufacturer = clean(row[indexes["生产厂家"]], 160)
        if len(name) < 2 or not indication or not approval or not manufacturer:
            continue
        record = {
            "name": name,
            "specification": clean(row[indexes["包装规格"]], 240),
            "approvalNumber": approval,
            "manufacturer": manufacturer,
            "category": clean(row[indexes["子分类"]], 80),
            "url": clean(row[indexes["详情链接"]], 500),
            "indication": indication,
            "usage": clean(row[indexes["用法用量"]], 700),
            "adverseReaction": clean(row[indexes["不良反应"]], 500),
            "contraindication": clean(row[indexes["禁忌"]], 500),
            "precaution": clean(row[indexes["注意事项"]], 500),
            "children": clean(row[indexes["儿童用药"]], 350),
            "elderly": clean(row[indexes["老年患者用药"]], 350),
            "pregnancyLactation": clean(row[indexes["孕妇及哺乳期妇女用药"]], 350),
            "interaction": clean(row[indexes["药物相互作用"]], 500),
        }
        record["fingerprint"] = fingerprint(record)
        previous = entries.get(name)
        if previous is None or sum(map(len, record.values())) > sum(map(len, previous.values())):
            entries[name] = record

    payload = {
        "schemaVersion": "local-patent-medicine-label-index-v1",
        "sourceFile": source.name,
        "sourceSha256": file_sha256(source),
        "sourceRowCount": source_row_count,
        "selectionBoundary": "主分类=中成药 AND 药品类型=非处方药；按规范药名保留字段最完整的一条",
        "entryCount": len(entries),
        "entries": [entries[name] for name in sorted(entries)],
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Built {output}: {len(entries)} unique OTC Chinese patent medicines from {source_row_count} rows")


if __name__ == "__main__":
    main()
