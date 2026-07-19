#!/usr/bin/env python3
"""Build the compact, runtime formula-provenance catalog from the local workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = PROJECT_ROOT.parent / "合理用药/docker/offline/pharma_reference_data/中医方剂Excel数据表_84295.xlsx"
DEFAULT_CLASSIC_SOURCE = PROJECT_ROOT.parent / "合理用药/docker/offline/pharma_reference_data/古代经典名方（第一批、第二批汉族医药、第二篇儿科部分）汇总-20241210.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "src/data/tcm-formula-sources.json"
NUMBER = r"(?:\d+(?:\.\d+)?|[一二三四五六七八九十百千万]+|半)"
UNIT = r"(?:克|毫克|两|钱|分|厘|斤|升|合|枚|个|片|撮|匙|盏|束|字|g|mg)"
# Historical prescriptions frequently write 1.5 qian as “钱半” without the leading “一”. Treat
# that construction as dose syntax as well; otherwise it is silently appended to the herb name
# (e.g. 青蒿脑钱) and makes an official formula impossible to resolve against itself.
QUANTITY_RE = re.compile(
    rf"(?:{NUMBER}(?:至{NUMBER})?{UNIT}|钱(?:半|[一二三四五六七八九十]分)(?:至{NUMBER}钱)?|两半|数分(?:或)?)",
    re.I,
)
QUANTITY_RANGE_SEPARATOR_RE = re.compile(rf"({NUMBER})[、,]({NUMBER})(?={UNIT})", re.I)
DOSE_ONLY_RE = re.compile(rf"^(?:各)?(?:{NUMBER}(?:至{NUMBER})?{UNIT}|钱(?:半|[一二三四五六七八九十]分)|两半)(?:余|许|或)?$", re.I)
RELATIVE_DOSE_SUFFIX_RE = re.compile(
    r"(?:加(?:一|二|两|三|四|五|半|\d+(?:\.\d+)?)倍|如[^、，,；;。\n]{1,10}(?:大|小)|大者|数分或)$"
)
QUANTITY_COLLISION_INGREDIENTS = {"百合"}


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def ingredient_names(raw: object) -> list[str]:
    text = re.sub(r"\s+", "", str(raw or "").strip())
    text = re.sub(r"[（(][^）)]*[）)]", "", text)
    text = QUANTITY_RANGE_SEPARATOR_RE.sub(r"\1至\2", text)
    result: list[str] = []
    for item in re.split(r"[、，,；;。\n]+", text):
        clean = item.strip().lstrip("各")
        if re.match(r"^(?:或加至|一云|又云|另云)", clean):
            continue
        clean = RELATIVE_DOSE_SUFFIX_RE.sub("", clean).strip()
        quantity = next((match for match in QUANTITY_RE.finditer(clean) if match.start() > 0), None)
        if quantity:
            clean = clean[: quantity.start()]
        clean = RELATIVE_DOSE_SUFFIX_RE.sub("", clean).strip()
        clean = clean.rstrip("各")
        clean = re.sub(r"(?:各等分|等分|适量|少许|若干|许)$", "", clean).strip()
        clean = re.sub(r"^(?:上药|右药|以上)", "", clean).strip()
        dose_only = bool(DOSE_ONLY_RE.fullmatch(clean)) and clean not in QUANTITY_COLLISION_INGREDIENTS
        if 1 < len(clean) <= 20 and not dose_only and re.search(r"[\u4e00-\u9fff]", clean) and clean not in result:
            result.append(clean)
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=Path(os.environ.get("TCM_FORMULA_SOURCE_XLSX", DEFAULT_SOURCE)))
    parser.add_argument("--classic-source", type=Path, default=Path(os.environ.get("TCM_CLASSIC_FORMULA_SOURCE_XLSX", DEFAULT_CLASSIC_SOURCE)))
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    source = args.source.resolve()
    classic_source = args.classic_source.resolve()
    output = args.output.resolve()
    if not source.is_file():
        raise SystemExit(f"Formula workbook not found: {source}")
    if not classic_source.is_file():
        raise SystemExit(f"Official classic formula workbook not found: {classic_source}")

    worksheet = load_workbook(source, read_only=True, data_only=True).active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    required = ["名称", "配方", "出处"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise SystemExit(f"Formula workbook missing columns: {', '.join(missing)}")
    indexes = {name: headers.index(name) for name in required}

    formulas: dict[str, list[dict[str, object]]] = {}
    source_row_count = 0
    for row in rows:
        source_row_count += 1
        name = str(row[indexes["名称"]] or "").strip()
        source_text = str(row[indexes["出处"]] or "").strip()
        ingredients = ingredient_names(row[indexes["配方"]])
        if not name or not source_text or not ingredients:
            continue
        variant = {"source": source_text, "ingredients": ingredients}
        bucket = formulas.setdefault(name, [])
        if variant not in bucket:
            bucket.append(variant)

    classic_sheet = load_workbook(classic_source, read_only=True, data_only=True).active
    official_classics: dict[str, dict[str, object]] = {}
    for row in classic_sheet.iter_rows(min_row=3, values_only=True):
        name = str(row[1] or "").strip()
        original = str(row[2] or "").strip()
        prescription = str(row[3] or "").strip()
        if not name or not original or not prescription:
            continue
        source_match = re.search(r"《[^》]{2,80}》", original)
        if not source_match:
            continue
        official_classics[name] = {
            "source": source_match.group(0),
            "sourceOriginal": original,
            "prescription": prescription,
            "ingredients": ingredient_names(prescription),
            "dosageForm": str(row[5] or "").strip(),
            "catalogBatch": str(row[6] or "").strip(),
        }

    payload = {
        "schemaVersion": "tcm-formula-provenance-v2",
        "sourceFile": source.name,
        "sourceSha256": file_sha256(source),
        "sourceRowCount": source_row_count,
        "formulaNameCount": len(formulas),
        "sourceColumns": required,
        "officialClassicSourceFile": classic_source.name,
        "officialClassicSourceSha256": file_sha256(classic_source),
        "officialClassicFormulaCount": len(official_classics),
        "officialClassicFormulas": dict(sorted(official_classics.items())),
        "formulas": dict(sorted(formulas.items())),
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Built {output}: {len(formulas)} formula names from {source_row_count} rows; {len(official_classics)} official classics")


if __name__ == "__main__":
    main()
