#!/usr/bin/env python3
"""Build the governed classic-formula indication index used by M03 retrieval."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_CANDIDATES = (
    PROJECT_ROOT / "药学基础数据/中医药数据/中医方剂Excel数据表_84295.xlsx",
    PROJECT_ROOT.parent / "合理用药/docker/offline/pharma_reference_data/中医方剂Excel数据表_84295.xlsx",
)
FORMULA_SOURCE = PROJECT_ROOT / "src/data/tcm-formula-sources.json"
VERIFIED_SUPPLEMENT_SOURCE = PROJECT_ROOT / "src/data/tcm-verified-formula-supplements.json"
DEFAULT_OUTPUT = PROJECT_ROOT / "src/data/tcm-formula-indications.json"


def default_source() -> Path:
    configured = os.environ.get("TCM_FORMULA_SOURCE_XLSX")
    if configured:
        return Path(configured)
    return next((path for path in DEFAULT_SOURCE_CANDIDATES if path.is_file()), DEFAULT_SOURCE_CANDIDATES[0])


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalized_source(value: object) -> str:
    return re.sub(r"[\s。；;，,：:（）()]", "", str(value or ""))


def source_matches(governed: str, workbook_source: str) -> bool:
    left = normalized_source(governed)
    right = normalized_source(workbook_source)
    if not left or not right:
        return False
    return left in right or right in left


def stable_id(name: str, source: str) -> str:
    suffix = hashlib.sha256(f"{name}\0{source}".encode()).hexdigest()[:12]
    return f"TCM-FORMULA-{suffix.upper()}"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=default_source())
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    source = args.source.resolve()
    output = args.output.resolve()
    if not source.is_file():
        raise SystemExit(f"Formula workbook not found: {source}")

    formula_catalog = json.loads(FORMULA_SOURCE.read_text(encoding="utf-8"))
    verified_catalog = json.loads(VERIFIED_SUPPLEMENT_SOURCE.read_text(encoding="utf-8"))
    governed: dict[str, dict[str, object]] = {}
    for name, entry in formula_catalog.get("officialClassicFormulas", {}).items():
        governed[name] = {
            "name": name,
            "source": str(entry.get("source") or "").strip(),
            "ingredients": entry.get("ingredients") or [],
            "catalog": "official_classic_catalog",
        }
    for name, entry in verified_catalog.get("entries", {}).items():
        governed.setdefault(name, {
            "name": name,
            "source": str(entry.get("source") or "").strip(),
            "ingredients": entry.get("ingredients") or [],
            "catalog": "verified_reference_catalog",
        })
        governed[name]["curatedIndications"] = [
            str(value).strip()
            for value in entry.get("indications", [])
            if str(value).strip()
        ]

    worksheet = load_workbook(source, read_only=True, data_only=True).active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    required = ["名称", "出处", "功效"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise SystemExit(f"Formula workbook missing columns: {', '.join(missing)}")
    indexes = {name: headers.index(name) for name in required}

    matches: dict[str, list[str]] = {name: [] for name in governed}
    source_row_count = 0
    for row in rows:
        source_row_count += 1
        name = str(row[indexes["名称"]] or "").strip()
        entry = governed.get(name)
        if not entry:
            continue
        workbook_source = str(row[indexes["出处"]] or "").strip()
        indication = re.sub(r"\s+", "", str(row[indexes["功效"]] or "").strip())
        if not indication or not source_matches(str(entry["source"]), workbook_source):
            continue
        bucket = matches[name]
        if indication not in bucket:
            bucket.append(indication)

    entries = []
    for name, entry in sorted(governed.items()):
        indications = list(entry.pop("curatedIndications", []))
        for indication in matches[name]:
            if indication not in indications:
                indications.append(indication)
        if not indications:
            continue
        entries.append({
            "id": stable_id(name, str(entry["source"])),
            **entry,
            "indications": indications[:3],
        })

    payload = {
        "schemaVersion": "tcm-formula-indications-v1",
        "sourceFile": source.name,
        "sourceSha256": file_sha256(source),
        "sourceRowCount": source_row_count,
        "governedFormulaCount": len(governed),
        "matchedFormulaCount": len(entries),
        "entries": entries,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Built {output}: {len(entries)}/{len(governed)} governed formulas have source-matched indications")


if __name__ == "__main__":
    main()
