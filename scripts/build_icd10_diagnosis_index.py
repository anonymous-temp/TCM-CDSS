#!/usr/bin/env python3
"""Build a compact deterministic ICD-10 lookup index from the supplied payer workbook."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "药学基础数据/药学相关资料/ICD-10医保2.0版-中英对应-20220426.xlsx"
OUTPUT = ROOT / "src/data/icd10-diagnosis-index.json"


def clean(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def aliases(name: str) -> set[str]:
    values = {name}
    for match in re.finditer(r"[\[［【(（]([^\]］】)）]{2,40})[\]］】)）]", name):
        values.add(match.group(1))
        outside = (name[: match.start()] + name[match.end() :]).strip()
        if outside:
            values.add(outside)
    return {item for item in values if len(item) >= 2}


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    entries: dict[tuple[str, str, str], dict[str, object]] = {}
    columns = (
        (13, 14, "diagnosis"),
        (10, 11, "subcategory"),
        (7, 8, "category"),
    )
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for code_index, name_index, level in columns:
            code = clean(row[code_index])
            name = clean(row[name_index])
            if not code or not name or name == "#N/A" or not re.match(r"^[A-Z]\d{2}", code):
                continue
            key = (code, name, level)
            record = entries.setdefault(key, {"code": code, "name": name, "level": level, "aliases": set()})
            record["aliases"].update(aliases(name))

    payload_entries = []
    for record in entries.values():
        payload_entries.append({
            "code": record["code"],
            "name": record["name"],
            "level": record["level"],
            "aliases": sorted(record["aliases"]),
        })
    level_order = {"diagnosis": 0, "subcategory": 1, "category": 2}
    payload_entries.sort(key=lambda item: (level_order[item["level"]], item["code"], item["name"]))
    payload = {
        "schemaVersion": "tcm-cdss-icd10-index-v1",
        "source": SOURCE.name,
        "entryCount": len(payload_entries),
        "entries": payload_entries,
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({"source": str(SOURCE), "output": str(OUTPUT), "entries": len(payload_entries)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
